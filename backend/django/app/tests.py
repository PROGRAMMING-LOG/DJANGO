import traceback

from django.test import TestCase

# Selenium
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.by import By

# Openpyxl
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side

import regex

from django.conf import settings
from django.contrib.auth.models import User


# Logger
from package.logger import Logger
log = Logger('/django/app/tests.py')

# log.debug('PASS')


# #########
# Selenium
# #########
chrome_options = Options()
desired_capabilities = chrome_options.to_capabilities()

driver = webdriver.Remote(
    command_executor='http://selenium:4444/wd/hub',
    desired_capabilities=desired_capabilities
)

url = 'http://nginx:8000/'

driver.get(url)
log.debug(driver.current_url)

# なにかの処理
try:
    # 要素が見つかるまで30秒待機
    log.debug('WAIT')
    WebDriverWait(driver, 30).until(
        expected_conditions.presence_of_element_located((By.TAG_NAME, 'h1'))
    )
except:
    # 要素が見つからなかったら終了
    driver.close()
    driver.quit()
    traceback.print_exc()

# 要素が見つかったら要素内のテキストを取得
text = driver.find_element(By.TAG_NAME, 'h1').text
log.debug(text)

driver.close()
driver.quit()


# #########
# Openpyxl
# #########
def is_ja(string):
    """
    日本語判定
    「ー」が記号扱いのようで拾えないが、「ー」のみの文章は有用性のないものと定義してスルーする
    """
    # 漢字
    if regex.search(r'\p{Han}+', string):
        return True

    # ひらがな
    if regex.search(r'\p{Hiragana}+', string):
        return True

    # カタカナ
    if regex.search(r'\p{Katakana}+', string):
        return True

    return False


wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'リスト'

# フォント
font_black = openpyxl.styles.Font(
    name='游ゴシック',
    color='000000',
    size=10,
)
font_black_hyperlink = openpyxl.styles.Font(
    name='游ゴシック',
    color='000000',
    size=11,
    underline='single',
)
font_white = openpyxl.styles.Font(
    name='游ゴシック',
    color='FFFFFF',
    size=10,
    bold=True,
)
font_repetition = openpyxl.styles.Font(
    name='游ゴシック',
    color='FF0000',
    size=10,
)
font_repetition_hyperlink = openpyxl.styles.Font(
    name='游ゴシック',
    color='DC143C',
    size=11,
    underline='single',
)

# セル背景色
fill_black = PatternFill(fill_type='solid', fgColor='000000')
fill_gray = PatternFill(fill_type='solid', fgColor='D9D9D9')
fill_yellow = PatternFill(fill_type='solid', fgColor='FFFF00')

# 整列
cell_center = Alignment(
    horizontal='center',
    vertical='center'
)
cell_left = Alignment(
    horizontal='left',
    vertical='center'
)

# 枠線
cell_border = Border(
    left=Side(
        border_style='thin',
        color='000000'
    ),
    right=Side(
        border_style='thin',
        color='000000'
    ),
    top=Side(
        border_style='thin',
        color='000000'
    ),
    bottom=Side(
        border_style='thin',
        color='000000'
    )
)

# タイトル
content_data = {
    'B2': {
        'value': 'ユーザー名',
    },
    'C2': {
        'value': 'メールアドレス',
    },
}

for cell in content_data:
    # 現在のセル
    current_cell = ws[cell]
    current_cell.value = content_data[cell]['value']
    current_cell.font = font_black
    current_cell.fill = fill_gray
    current_cell.alignment = cell_center
    current_cell.border = cell_border

# #########
# コンテンツセット
# #########

for i, user in enumerate(User.objects.all()):
    for cell in content_data:
        # アルファベット
        cell_alphabet = regex.sub(r'[0-9]+', '', cell)
        # 数字
        cell_number = regex.sub(r'[a-zA-Z]+', '', cell)

        current_number = str(int(cell_number) + (int(i) + 1))

        # 現在のセル
        current_cell = ws[cell_alphabet + current_number]
        current_cell.font = font_black
        current_cell.alignment = cell_left
        current_cell.border = cell_border

        value = ''
        if content_data[cell]['value'] == 'ユーザー名':
            value = user.username

        elif content_data[cell]['value'] == 'メールアドレス':
            value = user.email

        ws[cell_alphabet + current_number] = value

# セル幅 基本高
for row in ws.rows:
    first_row = row[0].row
    ws.row_dimensions[first_row].height = 18

# セル幅 自動伸縮
for col in ws.columns:
    max_length = 0
    empty = True
    column = openpyxl.utils.get_column_letter(col[0].column)

    # log.debug(column)

    for cell in col:
        base_length = 0
        empty = True

        if cell.value:
            empty = False

        # log.debug(str(cell.value))
        # log.debug(len(str(cell.value)))

        if empty is False and (len(str(cell.value)) >= base_length):
            ratio = 1.2

            # 小文字のみの場合は倍率を調整する
            if not is_ja(str(cell.value)) and not regex.search(r'[A-Z]', str(cell.value)):
                ratio = 0.95

            base_length = len(str(cell.value)) * ratio

            if base_length > max_length:
                max_length = base_length

                if max_length < 24:
                    max_length = 24

    ws.column_dimensions[column].width = max_length

wb.save(settings.BASE_DIR + '/app/' + 'data.xlsx')
